#!/usr/bin/env python3
"""
ingest.py - Scan the Public vault, extract text, generate embeddings via Ollama,
and store them in a local ChromaDB instance.

Usage:
    source /Users/chaomingou/Documents/VaultTools/.venv/bin/activate
    python /Users/chaomingou/Documents/VaultTools/ingest.py
"""

import os
import hashlib
import json
import subprocess
import re
import chromadb
from pathlib import Path
from typing import Optional, List

# ─── Configuration ───────────────────────────────────────────────────────────

VAULT_PATH = Path("/Users/chaomingou/Public")
CHROMA_PATH = Path("/Users/chaomingou/Documents/VaultTools/.chromadb")
COLLECTION_NAME = "public_vault"
OLLAMA_EMBED_MODEL = "nomic-embed-text"

# File types we support
SUPPORTED_EXTENSIONS = {".md", ".txt", ".pdf", ".xlsx", ".csv"}

# Directories to skip
SKIP_DIRS = {".venv", ".obsidian", ".chromadb", ".git", "node_modules", "99_Templates"}

# Chunk size (characters)
CHUNK_SIZE = 800
CHUNK_OVERLAP = 200


# ─── Text Extraction ────────────────────────────────────────────────────────

def extract_text_from_md(filepath: Path) -> str:
    """Read markdown/text files."""
    try:
        return filepath.read_text(encoding="utf-8", errors="ignore")
    except Exception as e:
        print(f"  ⚠ Error reading {filepath.name}: {e}")
        return ""


def extract_text_from_pdf(filepath: Path) -> str:
    """Extract text from PDF using pypdf."""
    try:
        from pypdf import PdfReader
        reader = PdfReader(str(filepath))
        text = ""
        for page in reader.pages:
            page_text = page.extract_text()
            if page_text:
                text += page_text + "\n"
        return text
    except Exception as e:
        print(f"  ⚠ Error reading PDF {filepath.name}: {e}")
        return ""


def extract_text_from_xlsx(filepath: Path) -> str:
    """Extract text from Excel files."""
    try:
        from openpyxl import load_workbook
        wb = load_workbook(str(filepath), read_only=True, data_only=True)
        text_parts = []
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            text_parts.append(f"--- Sheet: {sheet} ---")
            for row in ws.iter_rows(values_only=True):
                row_text = " | ".join(str(c) for c in row if c is not None)
                if row_text.strip():
                    text_parts.append(row_text)
        wb.close()
        return "\n".join(text_parts)
    except Exception as e:
        print(f"  ⚠ Error reading Excel {filepath.name}: {e}")
        return ""


def extract_text_from_csv(filepath: Path) -> str:
    """Read CSV files as text."""
    try:
        return filepath.read_text(encoding="utf-8", errors="ignore")
    except Exception as e:
        print(f"  ⚠ Error reading CSV {filepath.name}: {e}")
        return ""


def extract_text(filepath: Path) -> str:
    """Route file to the correct extractor."""
    ext = filepath.suffix.lower()
    if ext in {".md", ".txt"}:
        return extract_text_from_md(filepath)
    elif ext == ".pdf":
        return extract_text_from_pdf(filepath)
    elif ext == ".xlsx":
        return extract_text_from_xlsx(filepath)
    elif ext == ".csv":
        return extract_text_from_csv(filepath)
    return ""


# ─── Chunking ────────────────────────────────────────────────────────────────

def chunk_text(text: str, chunk_size: int = CHUNK_SIZE, overlap: int = CHUNK_OVERLAP) -> List[str]:
    """Split text into overlapping chunks."""
    if not text.strip():
        return []
    
    chunks = []
    start = 0
    while start < len(text):
        end = start + chunk_size
        chunk = text[start:end]
        if chunk.strip():
            chunks.append(chunk.strip())
        start += chunk_size - overlap
    return chunks


# ─── Embedding via Ollama ────────────────────────────────────────────────────

def get_embedding(text: str) -> Optional[List[float]]:
    """Get embedding vector from Ollama."""
    # Prefix for nomic-embed-text
    prefixed_text = f"search_document: {text}"
    try:
        # Try the API approach instead
        import urllib.request
        data = json.dumps({"model": OLLAMA_EMBED_MODEL, "input": prefixed_text[:2000]}).encode()
        req = urllib.request.Request(
            "http://localhost:11434/api/embed",
            data=data,
            headers={"Content-Type": "application/json"}
        )
        with urllib.request.urlopen(req, timeout=30) as resp:
            result = json.loads(resp.read().decode())
            embeddings = result.get("embeddings")
            if embeddings and len(embeddings) > 0:
                return embeddings[0]
        return None
    except Exception as e:
        print(f"  ⚠ Embedding error: {e}")
        return None


# ─── File Discovery ──────────────────────────────────────────────────────────

def discover_files(vault_path: Path) -> list[Path]:
    """Find all supported files in the vault."""
    files = []
    for root, dirs, filenames in os.walk(vault_path):
        # Skip hidden and excluded dirs
        dirs[:] = [d for d in dirs if d not in SKIP_DIRS and not d.startswith(".")]
        
        for fname in filenames:
            fpath = Path(root) / fname
            if fpath.suffix.lower() in SUPPORTED_EXTENSIONS:
                files.append(fpath)
    return sorted(files)


def file_hash(filepath: Path) -> str:
    """Generate a hash for the file to detect changes."""
    stat = filepath.stat()
    return hashlib.md5(f"{filepath}:{stat.st_size}:{stat.st_mtime}".encode()).hexdigest()


# ─── Category Detection ─────────────────────────────────────────────────────

def detect_category(filepath: Path) -> str:
    """Detect the category based on the folder structure."""
    rel = filepath.relative_to(VAULT_PATH)
    parts = rel.parts
    if len(parts) > 1:
        return parts[0]  # e.g., "00_Career", "10_Finance"
    return "Root"


# ─── Main Ingestion ──────────────────────────────────────────────────────────

def main():
    import sys
    subpath = sys.argv[1] if len(sys.argv) > 1 else ""
    target_vault = VAULT_PATH / subpath
    
    print("🧠 Public Vault Ingestion")
    print(f"   Vault: {target_vault}")
    print(f"   ChromaDB: {CHROMA_PATH}")
    print()

    # Initialize ChromaDB
    client = chromadb.PersistentClient(path=str(CHROMA_PATH))
    
    # We'll stick with get_or_create to allow targeted indexing without wiping everything
    collection = client.get_or_create_collection(
        name=COLLECTION_NAME,
        metadata={"hnsw:space": "cosine"}
    )

    # Discover files
    files = discover_files(target_vault)
    print(f"📂 Found {len(files)} supported files\n")

    total_chunks = 0
    errors = 0

    for i, filepath in enumerate(files, 1):
        # Calculate category relative to VAULT_PATH even if targeting subpath
        category = detect_category(filepath)
        print(f"[{i}/{len(files)}] {category}/{filepath.name}")

        # Extract text
        text = extract_text(filepath)
        if not text.strip():
            print("  ⏭ No text extracted, skipping.")
            continue

        # Chunk the text
        chunks = chunk_text(text)
        if not chunks:
            print("  ⏭ No chunks generated, skipping.")
            continue

        print(f"  📝 {len(chunks)} chunks")

        for j, chunk in enumerate(chunks):
            chunk_id = f"{file_hash(filepath)}_{j}"
            
            # Get embedding
            embedding = get_embedding(chunk)
            if embedding is None:
                errors += 1
                continue

            # Store in ChromaDB
            collection.add(
                ids=[chunk_id],
                embeddings=[embedding],
                documents=[chunk],
                metadatas=[{
                    "source": str(filepath),
                    "filename": filepath.name,
                    "category": category,
                    "chunk_index": j,
                    "total_chunks": len(chunks),
                }]
            )
            total_chunks += 1

        print(f"  ✅ Stored {len(chunks)} chunks")

    print(f"\n{'='*50}")
    print(f"✅ Ingestion complete!")
    print(f"   Total files processed: {len(files)}")
    print(f"   Total chunks stored: {total_chunks}")
    print(f"   Errors: {errors}")
    print(f"   Database location: {CHROMA_PATH}")


if __name__ == "__main__":
    main()
